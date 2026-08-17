from openpyxl import load_workbook

from normalizer import (
    normalize_text,
    normalize_email,
    normalize_phone
)


def load_companies(file_path):
    """
    Load company records from an Excel file.
    """

    workbook = load_workbook(
        file_path,
        read_only=True,
        data_only=True
    )

    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        return []

    headers = [
        normalize_text(header)
        for header in rows[0]
    ]

    companies = []

    for row in rows[1:]:

        company = {}

        for index, header in enumerate(headers):

            if index < len(row):
                company[header] = row[index]

        companies.append(company)

    return companies
def find_company(company_name, companies):
    """
    Find a company by name.
    """

    target = normalize_text(company_name)

    for company in companies:

        stored_name = company.get("company name", "")

        if normalize_text(stored_name) == target:
            return company

    return None
def verify_company_details(
    company_name,
    address=None,
    email=None,
    phone=None,
    companies=None
):
    """
    Compare extracted company details
    against the verified company database.
    """

    if companies is None:
        companies = []

    company = find_company(company_name, companies)

    if not company:
        return {
            "company_found": False,
            "address_match": False,
            "email_match": False,
            "phone_match": False
        }

    stored_address = company.get("address")
    stored_email = company.get("email")
    stored_phone = company.get("phone")

    address_match = None
    email_match = None
    phone_match = None

    if address and stored_address:
        address_match = (
            normalize_text(address)
            == normalize_text(stored_address)
        )

    if email and stored_email:
        email_match = (
            normalize_email(email)
            == normalize_email(stored_email)
        )

    if phone and stored_phone:
        phone_match = (
            normalize_phone(phone)
            == normalize_phone(stored_phone)
        )

    return {
        "company_found": True,
        "address_match": address_match,
        "email_match": email_match,
        "phone_match": phone_match
    }